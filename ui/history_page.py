import os
import sqlite3

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QFrame, QTableWidget, QTableWidgetItem, QHeaderView,
    QCheckBox, QAbstractItemView, QMenu, QFileDialog,
    QMessageBox, QCalendarWidget, QDialog, QSpinBox
)
from PyQt6.QtCore import Qt, QDate, QTimer, QEvent, QSize
from PyQt6.QtGui import QColor, QCursor, QIcon

from utils.helpers import get_asset_path, format_number

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "database", "database.db")


# ─────────────────────────────────────────────────────────────────
#  Fixed Calendar Widget
# ─────────────────────────────────────────────────────────────────
class FixedCalendar(QCalendarWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setGridVisible(False)
        self.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)
        QTimer.singleShot(0, self._patch_year_spinbox)

    def _patch_year_spinbox(self):
        spinbox = self.findChild(QSpinBox, "qt_calendar_yearedit")
        if spinbox:
            spinbox.setKeyboardTracking(True)
            spinbox.editingFinished.connect(self._commit_year)
            spinbox.installEventFilter(self)

    def _commit_year(self):
        spinbox = self.findChild(QSpinBox, "qt_calendar_yearedit")
        if spinbox:
            year = spinbox.value()
            cur  = self.selectedDate()
            new_date = QDate(year, cur.month(), min(cur.day(), QDate(year, cur.month(), 1).daysInMonth()))
            self.setSelectedDate(new_date)
            self.setCurrentPage(new_date.year(), new_date.month())

    def eventFilter(self, obj, event):
        if isinstance(obj, QSpinBox) and event.type() == QEvent.Type.KeyPress:
            if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
                self._commit_year()
                return True
        return super().eventFilter(obj, event)


# ─────────────────────────────────────────────────────────────────
#  Date Range Picker Dialog
# ─────────────────────────────────────────────────────────────────
class DateRangeDialog(QDialog):
    def __init__(self, parent=None, start_date=None, end_date=None):
        super().__init__(parent)
        self.setWindowTitle("Date Filter")
        self.setModal(True)
        self.setFixedSize(660, 340)
        self._date_was_set = False
        self.setStyleSheet("""
            QDialog { background-color: #1e1f25; color: #e2e2e9; font-family: 'Segoe UI', Arial; }
            QLabel  { color: #e2e2e9; font-size: 12px; font-weight: bold; background: transparent; }
            QCalendarWidget { background-color: #282a2f; color: #e2e2e9; border-radius: 8px; }
            QCalendarWidget QAbstractItemView {
                background-color: #282a2f; color: #e2e2e9;
                selection-background-color: #3D8EF0; selection-color: white;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar { background-color: #1e1f25; }
            QCalendarWidget QToolButton {
                color: #e2e2e9; background: transparent; font-size: 13px; font-weight: bold;
            }
            QCalendarWidget QToolButton:hover { background-color: #37393f; border-radius: 4px; }
            QCalendarWidget QSpinBox {
                color: #e2e2e9; background-color: #282a2f; border: none;
                font-size: 13px; font-weight: bold;
            }
            QPushButton {
                background-color: #282a2f; color: #e2e2e9; border: 1px solid #414752;
                border-radius: 6px; padding: 7px 18px; font-size: 13px; font-weight: bold;
            }
            QPushButton:hover { background-color: #37393f; }
            #btn-apply { background-color: #3D8EF0; color: white; border: none; }
            #btn-apply:hover { background-color: #5b9ff2; }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        cal_row = QHBoxLayout()

        start_col = QVBoxLayout()
        start_col.addWidget(QLabel("Start date"))
        self.cal_start = FixedCalendar()
        if start_date:
            self.cal_start.setSelectedDate(start_date)
        self.cal_start.clicked.connect(lambda _: self._mark_set())
        start_col.addWidget(self.cal_start)

        end_col = QVBoxLayout()
        end_col.addWidget(QLabel("End date"))
        self.cal_end = FixedCalendar()
        if end_date:
            self.cal_end.setSelectedDate(end_date)
        self.cal_end.clicked.connect(lambda _: self._mark_set())
        end_col.addWidget(self.cal_end)

        cal_row.addLayout(start_col)
        cal_row.addSpacing(15)
        cal_row.addLayout(end_col)
        layout.addLayout(cal_row)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_reset = QPushButton("Reset")
        btn_reset.clicked.connect(self._reset)
        btn_apply = QPushButton("Set")
        btn_apply.setObjectName("btn-apply")
        btn_apply.clicked.connect(self.accept)
        btn_row.addWidget(btn_reset)
        btn_row.addSpacing(10)
        btn_row.addWidget(btn_apply)
        layout.addLayout(btn_row)

    def _mark_set(self):
        self._date_was_set = True

    def _reset(self):
        today = QDate.currentDate()
        self.cal_start.setSelectedDate(today.addMonths(-1))
        self.cal_end.setSelectedDate(today)
        self._date_was_set = False

    def get_dates(self):
        return self.cal_start.selectedDate(), self.cal_end.selectedDate(), self._date_was_set


# ─────────────────────────────────────────────────────────────────
#  History Page
# ─────────────────────────────────────────────────────────────────
class HistoryPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._filter_start: QDate | None = None
        self._filter_end:   QDate | None = None
        self._filter_class: str | None   = None
        self._data: list[dict] = []

        self._init_ui()
        self._apply_styles()
        self.load_data()

        self._refresh_timer = QTimer(self)
        self._refresh_timer.setInterval(3000)
        self._refresh_timer.timeout.connect(self._auto_refresh)
        self._refresh_timer.start()

    # ── UI Construction ──────────────────────────────────────────────
    def _init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Header
        header = QFrame()
        header.setObjectName("top-header")
        header.setFixedHeight(64)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(25, 0, 25, 0)

        title_lbl = QLabel("History Log")
        title_lbl.setStyleSheet("font-size: 22px; font-weight: 900; color: #e2e2e9;")

        self.date_btn = QPushButton("📅  All Date  ▼")
        self.date_btn.setObjectName("date-btn")
        self.date_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.date_btn.clicked.connect(self._open_date_filter)

        self.export_top_btn = QPushButton("⬇  Export Tables")
        self.export_top_btn.setObjectName("btn-primary")
        self.export_top_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.export_top_btn.clicked.connect(self._export_all)

        header_layout.addWidget(title_lbl)
        header_layout.addSpacing(25)
        header_layout.addWidget(self.date_btn)
        header_layout.addStretch()
        header_layout.addWidget(self.export_top_btn)
        main_layout.addWidget(header)

        # Content
        content_wrapper = QWidget()
        content_layout = QVBoxLayout(content_wrapper)
        content_layout.setContentsMargins(30, 20, 30, 20)
        content_layout.setSpacing(25)

        # Stats card
        stats_frame = QFrame()
        stats_frame.setObjectName("stats-card")
        stats_frame.setFixedSize(280, 120)
        stats_layout = QVBoxLayout(stats_frame)
        stats_layout.setContentsMargins(25, 20, 25, 20)
        stats_title = QLabel("TOTAL COUNT")
        stats_title.setStyleSheet("font-size: 11px; font-weight: 800; color: #8b919e; letter-spacing: 1.5px;")
        self.stats_value = QLabel("0")
        self.stats_value.setStyleSheet(
            "font-size: 44px; font-weight: bold; color: #35e192; letter-spacing: -1px;"
        )
        stats_layout.addWidget(stats_title)
        stats_layout.addStretch()
        stats_layout.addWidget(self.stats_value)

        stats_wrap = QHBoxLayout()
        stats_wrap.addWidget(stats_frame)
        stats_wrap.addStretch()

        # Table container
        table_container = QFrame()
        table_container.setObjectName("table-container")
        tc_layout = QVBoxLayout(table_container)
        tc_layout.setContentsMargins(0, 0, 0, 0)
        tc_layout.setSpacing(0)

        # Toolbar
        tb_header = QWidget()
        tb_header.setFixedHeight(65)
        tb_header.setObjectName("tb-header")
        tb_h_layout = QHBoxLayout(tb_header)
        tb_h_layout.setContentsMargins(30, 0, 30, 0)

        tab_container = QWidget()
        tab_layout = QVBoxLayout(tab_container)
        tab_layout.setContentsMargins(0, 20, 0, 0)
        tab_lbl = QLabel("ALL EVENTS")
        tab_lbl.setStyleSheet(
            "font-size: 11px; font-weight:bold; color:#e2e2e9;"
            " border-bottom: 2px solid #3D8EF0; padding-bottom: 18px;"
        )
        tab_layout.addWidget(tab_lbl)

        tb_h_layout.addWidget(tab_container)
        tb_h_layout.addStretch()

        self.btn_del_sel = QPushButton("  Delete Selected")
        self.btn_del_sel.setIcon(QIcon(get_asset_path("trash.svg")))
        self.btn_del_sel.setObjectName("btn-danger-outline")
        self.btn_del_sel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_del_sel.clicked.connect(self._delete_selected)

        self.btn_exp_sel = QPushButton("⬇  Export Selected")
        self.btn_exp_sel.setObjectName("btn-outline")
        self.btn_exp_sel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_exp_sel.clicked.connect(self._export_selected)

        self.btn_filter_class = QPushButton("⚲  Filter by Class")
        self.btn_filter_class.setObjectName("btn-text")
        self.btn_filter_class.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_filter_class.clicked.connect(self._open_class_filter)

        tb_h_layout.addWidget(self.btn_del_sel)
        tb_h_layout.addSpacing(10)
        tb_h_layout.addWidget(self.btn_exp_sel)
        tb_h_layout.addSpacing(20)
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        sep.setStyleSheet("color: #33353a;")
        tb_h_layout.addWidget(sep)
        tb_h_layout.addSpacing(20)
        tb_h_layout.addWidget(self.btn_filter_class)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        headers = ["", "  DATE", "CLASS", "TOTAL COUNT", "ACTIONS"]
        for i, text in enumerate(headers):
            item = QTableWidgetItem(text)
            alignment = (
                Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
                if i == 1 else Qt.AlignmentFlag.AlignCenter
            )
            item.setTextAlignment(alignment)
            self.table.setHorizontalHeaderItem(i, item)

        self.header_chk = QCheckBox(self.table.horizontalHeader())
        self.header_chk.setStyleSheet("background: transparent;")
        self.header_chk.setGeometry(17, 12, 16, 16)
        self.header_chk.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.header_chk.stateChanged.connect(self._toggle_all_checkboxes)

        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 54)
        self.table.setColumnWidth(4, 90)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setShowGrid(False)
        self.table.verticalHeader().setVisible(False)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.setAlternatingRowColors(True)

        self._checkboxes: list[QCheckBox] = []

        tc_layout.addWidget(tb_header)
        tc_layout.addWidget(self.table)

        content_layout.addLayout(stats_wrap)
        content_layout.addWidget(table_container)
        main_layout.addWidget(content_wrapper, stretch=1)

    # ── Database — Read ──────────────────────────────────────────────
    def _fetch_from_db(self) -> list[dict]:
        if not os.path.exists(DB_PATH):
            return []
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        query = """
            SELECT DATE(timestamp) AS date, class_name, SUM(count_in) AS total_count
            FROM history_logs WHERE 1=1
        """
        params = []
        if self._filter_start and self._filter_end:
            query += " AND DATE(timestamp) BETWEEN ? AND ?"
            params += [
                self._filter_start.toString("yyyy-MM-dd"),
                self._filter_end.toString("yyyy-MM-dd"),
            ]
        if self._filter_class:
            query += " AND class_name = ?"
            params.append(self._filter_class)
        query += " GROUP BY DATE(timestamp), class_name ORDER BY date DESC, class_name"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        return [{"date": r[0], "class_name": r[1], "total_count": r[2]} for r in rows]

    def _fetch_classes(self) -> list[str]:
        if not os.path.exists(DB_PATH):
            return []
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT class_name FROM history_logs ORDER BY class_name")
        rows = [r[0] for r in cursor.fetchall()]
        conn.close()
        return rows

    # ── Database — Delete ────────────────────────────────────────────
    def _delete_rows_from_db(self, rows: list[dict]):
        if not rows or not os.path.exists(DB_PATH):
            return
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        for row in rows:
            cursor.execute(
                "DELETE FROM history_logs WHERE DATE(timestamp) = ? AND class_name = ?",
                (row["date"], row["class_name"])
            )
        conn.commit()
        conn.close()

    # ── Load / Refresh ───────────────────────────────────────────────
    def load_data(self):
        self._data = self._fetch_from_db()
        self._rebuild_table()

    def _auto_refresh(self):
        new_data = self._fetch_from_db()
        if new_data == self._data:
            return
        checked_keys = {
            (self._data[i]["date"], self._data[i]["class_name"])
            for i, chk in enumerate(self._checkboxes)
            if chk.isChecked() and i < len(self._data)
        }
        self._data = new_data
        self._rebuild_table(checked_keys=checked_keys)

    def _rebuild_table(self, checked_keys: set | None = None):
        self.table.setRowCount(0)
        self._checkboxes.clear()
        self.header_chk.setChecked(False)

        total_all = sum(r["total_count"] for r in self._data)
        self.stats_value.setText(format_number(total_all))

        for row_idx, row_data in enumerate(self._data):
            self.table.insertRow(row_idx)
            self.table.setRowHeight(row_idx, 52)

            # Col 0 — Checkbox
            chk_widget = QWidget()
            chk_layout = QHBoxLayout(chk_widget)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chk = QCheckBox()
            chk.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            if checked_keys and (row_data["date"], row_data["class_name"]) in checked_keys:
                chk.setChecked(True)
            chk.stateChanged.connect(lambda state, ri=row_idx: self._on_row_check(state, ri))
            chk_layout.addWidget(chk)
            self._checkboxes.append(chk)
            self.table.setCellWidget(row_idx, 0, chk_widget)
            if chk.isChecked():
                self._highlight_row(row_idx, True)

            # Col 1 — Date
            try:
                from datetime import datetime as _dt
                date_str = _dt.strptime(row_data["date"], "%Y-%m-%d").strftime("%d %B %Y")
            except Exception:
                date_str = row_data["date"]
            item_date = QTableWidgetItem(f"  {date_str}")
            item_date.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
            item_date.setFlags(item_date.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 1, item_date)

            # Col 2 — Class badge
            cls_widget = QWidget()
            cls_layout = QHBoxLayout(cls_widget)
            cls_layout.setContentsMargins(0, 0, 0, 0)
            cls_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            cls_lbl = QLabel()
            cls_lbl.setObjectName("class-badge")
            cls_lbl.setText(f'<span style="color:#4292f4;">●</span>  {row_data["class_name"]}')
            cls_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            cls_layout.addWidget(cls_lbl)
            self.table.setCellWidget(row_idx, 2, cls_widget)

            # Col 3 — Count
            count_item = QTableWidgetItem(format_number(row_data["total_count"]))
            count_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            count_item.setForeground(QColor("#35e192"))
            count_item.setFlags(count_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            f = count_item.font()
            f.setBold(True)
            f.setPointSize(12)
            count_item.setFont(f)
            self.table.setItem(row_idx, 3, count_item)

            # Col 4 — Delete button
            act_widget = QWidget()
            act_layout = QHBoxLayout(act_widget)
            act_layout.setContentsMargins(0, 0, 0, 0)
            act_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            btn_del = QPushButton()
            btn_del.setIcon(QIcon(get_asset_path("trash.svg")))
            btn_del.setIconSize(QSize(20, 20))
            btn_del.setObjectName("btn-action-delete")
            btn_del.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn_del.setFixedSize(32, 32)
            btn_del.clicked.connect(lambda _, rd=row_data: self._delete_single_row(rd))
            act_layout.addWidget(btn_del)
            self.table.setCellWidget(row_idx, 4, act_widget)

    # ── Row highlight ────────────────────────────────────────────────
    def _on_row_check(self, state, row_idx: int):
        self._highlight_row(row_idx, state != 0)

    def _highlight_row(self, row_idx: int, highlight: bool):
        bg = QColor("#1d3150") if highlight else QColor(0, 0, 0, 0)
        for col in [1, 3]:
            item = self.table.item(row_idx, col)
            if item:
                if highlight:
                    item.setBackground(bg)
                else:
                    item.setData(Qt.ItemDataRole.BackgroundRole, None)

    # ── Actions ──────────────────────────────────────────────────────
    def _delete_single_row(self, row_data: dict):
        reply = QMessageBox.question(
            self, "Konfirmasi Hapus",
            f"Hapus semua data untuk <b>{row_data['class_name']}</b> "
            f"pada <b>{row_data['date']}</b>?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._delete_rows_from_db([row_data])
            self.load_data()

    def _get_checked_rows(self) -> list[dict]:
        return [
            self._data[i]
            for i, chk in enumerate(self._checkboxes)
            if chk.isChecked() and i < len(self._data)
        ]

    def _delete_selected(self):
        rows = self._get_checked_rows()
        if not rows:
            QMessageBox.information(self, "Tidak Ada Pilihan", "Pilih baris terlebih dahulu.")
            return
        reply = QMessageBox.question(
            self, "Konfirmasi Hapus",
            f"Hapus {len(rows)} baris data yang dipilih?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._delete_rows_from_db(rows)
            self.load_data()

    # ── Export ───────────────────────────────────────────────────────
    def _export_rows_to_excel(self, rows: list[dict]):
        if not rows:
            QMessageBox.information(self, "Tidak Ada Data", "Tidak ada data untuk diekspor.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Simpan Excel", "history_export.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "History Log"

            header_fill   = PatternFill("solid", fgColor="1E3A5F")
            header_font   = Font(bold=True, color="FFFFFF", size=11, name="Segoe UI")
            header_align  = Alignment(horizontal="center", vertical="center")
            header_border = Border(bottom=Side(style="thin", color="3D8EF0"))
            col_headers   = ["No", "Date", "Class", "Total Count"]
            col_widths    = [6, 20, 22, 16]

            for col_idx, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = header_align
                cell.border    = header_border
                ws.column_dimensions[get_column_letter(col_idx)].width = w
            ws.row_dimensions[1].height = 24

            count_font   = Font(name="Segoe UI", bold=True, size=10, color="000000")
            base_font    = Font(name="Segoe UI", size=10, color="000000")
            center_align = Alignment(horizontal="center", vertical="center")
            left_align   = Alignment(horizontal="left",   vertical="center")
            white_fill   = PatternFill("solid", fgColor="FFFFFF")

            for i, row in enumerate(rows):
                r = i + 2
                try:
                    from datetime import datetime as _dt
                    date_str = _dt.strptime(row["date"], "%Y-%m-%d").strftime("%d %B %Y")
                except Exception:
                    date_str = row["date"]
                
                c1 = ws.cell(r, 1, i + 1)
                c1.alignment = center_align
                c1.font = base_font
                c1.fill = white_fill
                
                c2 = ws.cell(r, 2, date_str)
                c2.alignment = left_align
                c2.font = base_font
                c2.fill = white_fill
                
                c3 = ws.cell(r, 3, row["class_name"])
                c3.alignment = center_align
                c3.font = base_font
                c3.fill = white_fill
                
                count_cell = ws.cell(r, 4, row["total_count"])
                count_cell.font      = count_font
                count_cell.alignment = center_align
                count_cell.fill      = white_fill
                
                ws.row_dimensions[r].height = 20

            ws.freeze_panes = "A2"
            wb.save(path)
            QMessageBox.information(self, "Berhasil", f"Data berhasil diekspor ke:\n{path}")
        except ImportError:
            QMessageBox.critical(
                self, "Library Tidak Ditemukan",
                "Modul 'openpyxl' belum terinstall.\n\nJalankan: pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal mengekspor: {e}")

    def _export_all(self):
        self._export_rows_to_excel(self._data)

    def _export_selected(self):
        rows = self._get_checked_rows()
        if not rows:
            QMessageBox.information(self, "Tidak Ada Pilihan", "Pilih baris terlebih dahulu.")
            return
        self._export_rows_to_excel(rows)

    # ── Filters ──────────────────────────────────────────────────────
    def _open_date_filter(self):
        dlg = DateRangeDialog(self, self._filter_start, self._filter_end)
        if dlg.exec():
            s, e, was_set = dlg.get_dates()
            if not was_set:
                self._filter_start = None
                self._filter_end   = None
                self.date_btn.setText("📅  All Date  ▼")
            elif s <= e:
                self._filter_start = s
                self._filter_end   = e
                self.date_btn.setText(f"📅  {s.toString('dd MMM yyyy')} – {e.toString('dd MMM yyyy')}  ▼")
            else:
                QMessageBox.warning(self, "Tanggal Salah", "Tanggal mulai harus sebelum tanggal akhir.")
                return
            self.load_data()

    def _open_class_filter(self):
        classes = self._fetch_classes()
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: #1e1f25; color: #e2e2e9;
                border: 1px solid #414752; border-radius: 6px;
                padding: 6px; font-size: 13px;
            }
            QMenu::item { padding: 8px 20px; border-radius: 4px; }
            QMenu::item:selected { background-color: #3D8EF0; color: white; }
            QMenu::separator { height: 1px; background: #33353a; margin: 4px 0; }
        """)
        action_all = menu.addAction("✓ All Class" if not self._filter_class else "  All Class")
        action_all.setData(None)
        menu.addSeparator()
        for cls in classes:
            label = (f"✓ {cls}" if cls == self._filter_class else f"  {cls}")
            action = menu.addAction(label)
            action.setData(cls)

        pos    = self.btn_filter_class.mapToGlobal(self.btn_filter_class.rect().bottomLeft())
        chosen = menu.exec(pos)
        if chosen is not None:
            self._filter_class = chosen.data()
            self.btn_filter_class.setText(
                f"⚲  {self._filter_class}  ✕" if self._filter_class else "⚲  Filter by Class"
            )
            self.load_data()

    # ── Checkbox helpers ─────────────────────────────────────────────
    def _toggle_all_checkboxes(self, state):
        is_checked = (state != 0)
        for chk in self._checkboxes:
            if chk.isChecked() != is_checked:
                chk.setChecked(is_checked)

    # ── Styles ───────────────────────────────────────────────────────
    def _apply_styles(self):
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet("""
            HistoryPage { background-color: #111318; }
            #top-header { background-color: #111318; border-bottom: 1px solid #282a2f; }
            QWidget { color: #e2e2e9; font-family: 'Segoe UI', 'Inter', Arial, sans-serif; }
            #btn-primary {
                background-color: #3D8EF0; color: white; border-radius: 6px;
                padding: 8px 18px; font-weight: bold; font-size: 13px; border: none;
            }
            #btn-primary:hover { background-color: #5b9ff2; }
            #date-btn {
                background-color: #1a1b21; color: #8b919e; border: 1px solid #33353a;
                border-radius: 8px; padding: 6px 14px; font-size: 12px; font-weight: bold;
            }
            #date-btn:hover { background-color: #282a2f; color: #e2e2e9; }
            #stats-card { background-color: #0c0e13; border-radius: 16px; }
            #table-container {
                background-color: #1e1f25; border-radius: 12px; border: 1px solid #282a2f;
            }
            #tb-header { background-color: transparent; border-bottom: 1px solid #33353a; }
            #tb-header QWidget { background-color: transparent; }
            #btn-danger-outline {
                background-color: rgba(147, 0, 10, 0.15); color: #ffb4ab;
                border: 1px solid rgba(255, 180, 171, 0.2); border-radius: 6px;
                padding: 6px 14px; font-size: 12px; font-weight: 600;
            }
            #btn-danger-outline:hover { background-color: rgba(147, 0, 10, 0.3); }
            #btn-outline {
                background-color: #282a2f; color: #e2e2e9; border: 1px solid #414752;
                border-radius: 6px; padding: 6px 14px; font-size: 12px; font-weight: 600;
            }
            #btn-outline:hover { background-color: #37393f; }
            #btn-text { background-color: transparent; color: #8b919e; border: none;
                        font-size: 12px; font-weight: bold; }
            #btn-text:hover { color: #e2e2e9; }
            QTableWidget {
                background-color: transparent; border: none;
                gridline-color: transparent; color: #e2e2e9; font-size: 13px; outline: none;
            }
            QTableWidget::item { border-bottom: 1px solid rgba(255,255,255,0.03); padding: 0px 4px; }
            QTableWidget::item:selected { background-color: #1d3150; }
            QHeaderView::section {
                background-color: rgba(40,42,47,0.5); color: #8b919e;
                font-size: 11px; font-weight: bold; letter-spacing: 1px;
                padding: 12px 10px; border: none;
            }
            #class-badge {
                background-color: #282a2f; color: #e2e2e9; border-radius: 6px;
                padding: 5px 10px; font-size: 12px; font-weight: bold;
            }
            #btn-action-delete {
                background-color: rgba(147,0,10,0.2); color: #ffb4ab;
                border-radius: 6px; font-size: 16px; border: none;
            }
            #btn-action-delete:hover { background-color: rgba(200,30,30,0.8); color: white; }
            QCheckBox::indicator {
                width: 17px; height: 17px; border: 1.5px solid #414752;
                border-radius: 4px; background-color: #111318;
            }
            QCheckBox::indicator:checked { background-color: #3D8EF0; border: 1.5px solid #3D8EF0; }
            QCheckBox::indicator:hover { border-color: #3D8EF0; }
            QTableWidget::item:alternate { background-color: #17181e; }
            QScrollBar:vertical { border: none; background: #111318; width: 8px; margin: 0px; }
            QScrollBar::handle:vertical { background: #33353a; min-height: 20px; border-radius: 4px; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { border: none; background: none; height: 0px; }
        """)
